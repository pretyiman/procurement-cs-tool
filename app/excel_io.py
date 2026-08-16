"""Import a Comparative Statement-shaped Excel file into the DB.

Expected shape (see docs/data-model.md and CS.xlsx):
  - a header row whose first cell is "Ser"
  - the row directly below it holds supplier names, one per rate column
  - rate columns run from the "Rate Quoted by Firms..." header to (not
    including) the "Lowest" header
  - data rows follow, each starting with an integer serial number; the
    first row whose first cell is not a number ends the item list
  - "NQ" / blank / "-" in a rate cell means that supplier did not quote
"""

import re
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, List, Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from .models import Department, DocumentLabels, Item, ItemMaster, Quote, Supplier, TaxType, Tender, TenderStatus

if TYPE_CHECKING:
    from .award_engine import PurchaseProposal
    from .cs_engine import ComparativeStatement, ItemResult, PackageTotal

_TAX_PATTERN = re.compile(r"(\d+(?:\.\d+)?)\s*%\s*(GST|PST)", re.IGNORECASE)
_NQ_VALUES = {"NQ", "-", ""}


def _is_nq(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip().upper() in _NQ_VALUES:
        return True
    return False


def _parse_tax(rows) -> tuple:
    """Existing CS.xlsx-shaped files only ever label tax as "X% GST" - PST
    detection is here so a file this app exported for a PST tender can be
    re-imported correctly too (see export_cs_xlsx)."""
    for row in rows:
        for cell in row:
            if isinstance(cell, str):
                match = _TAX_PATTERN.search(cell)
                if match:
                    return float(match.group(1)), TaxType(match.group(2).upper())
    return 18.0, TaxType.GST


def _find_inquiry_no(rows, header_idx: int) -> str:
    for row in rows[:header_idx]:
        for cell in row:
            if isinstance(cell, str) and "inquiry" in cell.lower():
                return cell.strip()
    return "UNSPECIFIED"


def get_or_create_supplier(session: Session, name: str) -> tuple[Supplier, bool]:
    """Returns (supplier, created). Matching is case-insensitive (trimmed)
    so "M/s Awan Tech" and "m/s awan tech" resolve to the same row instead
    of creating a near-duplicate. Compared in Python, not via SQL lower():
    SQLite's built-in LOWER() only folds ASCII, so a description containing
    e.g. "Ø" would silently mismatch and re-insert a duplicate."""
    name = name.strip()
    needle = name.lower()
    existing = next((s for s in session.exec(select(Supplier)).all() if s.name.lower() == needle), None)
    if existing:
        return existing, False
    supplier = Supplier(name=name)
    session.add(supplier)
    session.flush()
    return supplier, True


def get_or_create_department(session: Session, name: str) -> tuple[Department, bool]:
    """Returns (department, created); case-insensitive match, see
    get_or_create_supplier for why it's done in Python."""
    name = name.strip()
    needle = name.lower()
    existing = next((d for d in session.exec(select(Department)).all() if d.name.lower() == needle), None)
    if existing:
        return existing, False
    department = Department(name=name)
    session.add(department)
    session.flush()
    return department, True


def get_or_create_item_master(
    session: Session, part_no: str, description: str, default_unit: str = ""
) -> tuple[ItemMaster, bool]:
    """Reuse a catalog row when (part_no, description) already matches one,
    case-insensitively (see docs/data-model.md - part_no alone isn't
    unique for "NIV" items; see get_or_create_supplier for why the compare
    is done in Python, not SQL). Returns (item_master, created)."""
    part_no = (part_no or "").strip()
    description = (description or "").strip()
    default_unit = (default_unit or "").strip()

    part_needle, desc_needle = part_no.lower(), description.lower()
    existing = next(
        (
            im
            for im in session.exec(select(ItemMaster)).all()
            if im.part_no.lower() == part_needle and im.description.lower() == desc_needle
        ),
        None,
    )
    if existing:
        if not existing.default_unit and default_unit:
            existing.default_unit = default_unit
            session.add(existing)
        return existing, False

    item_master = ItemMaster(part_no=part_no, description=description, default_unit=default_unit)
    session.add(item_master)
    session.flush()
    return item_master, True


def import_tender(path: Union[str, Path], session: Session) -> Tender:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(
        (i for i, row in enumerate(rows) if row and isinstance(row[0], str) and row[0].strip() == "Ser"),
        None,
    )
    if header_idx is None:
        raise ValueError("Could not find header row (expected a 'Ser' column)")

    header_row = rows[header_idx]
    subheader_row = rows[header_idx + 1]

    rate_start_col = None
    lowest_col = None
    for j, val in enumerate(header_row):
        if not isinstance(val, str):
            continue
        low = val.lower()
        if rate_start_col is None and "rate quoted" in low:
            rate_start_col = j
        if rate_start_col is not None and "lowest" in low:
            lowest_col = j
            break
    if rate_start_col is None or lowest_col is None:
        raise ValueError(
            "Could not locate supplier rate columns "
            "('Rate Quoted by Firms...' / 'Lowest' headers not found)"
        )

    supplier_cols = list(range(rate_start_col, lowest_col))
    supplier_names = [subheader_row[c] for c in supplier_cols]
    if any(not isinstance(n, str) or not n.strip() for n in supplier_names):
        raise ValueError("Expected a supplier name in the row below the header, under each rate column")

    inquiry_no = _find_inquiry_no(rows, header_idx)
    tax_percent, tax_type = _parse_tax(rows)

    tender = Tender(
        inquiry_no=inquiry_no, tax_type=tax_type, tax_percent=tax_percent, status=TenderStatus.draft
    )
    session.add(tender)
    session.flush()

    suppliers = [get_or_create_supplier(session, name)[0] for name in supplier_names]

    data_start = header_idx + 2
    started = False
    for row in rows[data_start:]:
        if not row or all(v is None for v in row):
            continue  # blank spacer row (e.g. between subheader and data)

        if not isinstance(row[0], (int, float)):
            if started:
                break  # reached the totals/summary section
            continue  # not at the data rows yet

        started = True
        item_master, _ = get_or_create_item_master(
            session,
            part_no=str(row[1]) if row[1] is not None else "",
            description=str(row[2]) if row[2] is not None else "",
            default_unit=str(row[3]) if row[3] is not None else "",
        )
        item = Item(
            tender_id=tender.id,
            item_master_id=item_master.id,
            ser=int(row[0]),
            qty=float(row[4]) if row[4] is not None else 0.0,
        )
        session.add(item)
        session.flush()

        for col, supplier in zip(supplier_cols, suppliers):
            raw = row[col]
            rate = None if _is_nq(raw) else float(raw)
            session.add(Quote(item_id=item.id, supplier_id=supplier.id, rate=rate))

    session.commit()
    session.refresh(tender)
    return tender


def _simple_list_workbook(title: str, headers: list, rows: list) -> bytes:
    """Shared shape for the plain catalog/list exports below: a bold title
    banner merged across the columns, bold headers, sized columns, no
    pricing/signature-block machinery - those belong to the CS exports."""
    wb = Workbook()
    ws = wb.active
    # Excel sheet names can't contain \ / ? * [ ] : and are capped at 31
    # chars - the banner cell below keeps the real, unsanitized title.
    sheet_name = re.sub(r'[\\/?*\[\]:]', "-", title)[:31]
    ws.title = sheet_name
    bold = Font(bold=True)

    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.font = bold

    row_idx = 4
    for row_values in rows:
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col, value=value)
        row_idx += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_item_catalog_xlsx(items: list[ItemMaster]) -> bytes:
    rows = [(im.part_no, im.description, im.default_unit) for im in items]
    return _simple_list_workbook("Item Catalog", ["Part No", "Description", "Unit"], rows)


def export_supplier_list_xlsx(suppliers: list[Supplier]) -> bytes:
    rows = [
        (s.name, s.contact_person or "", s.phone or "", s.email or "", s.address or "", s.tax_no or "")
        for s in suppliers
    ]
    return _simple_list_workbook(
        "Suppliers", ["Name", "Contact Person", "Phone", "Email", "Address", "Tax No"], rows
    )


def export_department_list_xlsx(departments: list[Department]) -> bytes:
    rows = [(d.name,) for d in departments]
    return _simple_list_workbook("Departments", ["Name"], rows)


def export_rfq_item_list_xlsx(tender: Tender, items: list[Item]) -> bytes:
    """Just the RFQ's own item list (Ser/Part No/Description/Unit/Qty) -
    no pricing, matching what the RFQ item page itself shows now that
    suppliers/rates were moved off it to Quote Entry."""
    title = f"RFQ Item List - {tender.inquiry_no}"
    rows = [
        (item.ser, item.item_master.part_no, item.item_master.description, item.item_master.default_unit, item.qty)
        for item in items
    ]
    return _simple_list_workbook(title, ["Ser", "Part No", "Description", "Unit", "Qty"], rows)


def export_cs_xlsx(cs: "ComparativeStatement", labels: DocumentLabels, custom_fields: dict = None) -> bytes:
    """Render a ComparativeStatement (app/cs_engine.py) as an .xlsx workbook
    shaped like the original CS.xlsx: Ser/Part No/Description/A-U/Qty, one
    rate column per supplier, Lowest Firm/Rate/Total Value, LPR/Inc-Dec%,
    then totals and a per-firm summary block. Deliberately shaped so the
    app's own import_tender() can re-parse it (see
    test_reexporting_and_reimporting_round_trips_correctly) - not just a
    one-way report."""
    custom_fields = custom_fields or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparative Statement"
    bold = Font(bold=True)
    header_align = Alignment(wrap_text=True, horizontal="center", vertical="center")

    suppliers = sorted(cs.suppliers_by_id.values(), key=lambda s: s.name)
    n = len(suppliers)

    rate_start_col = 6  # after Ser/Part No/Description/A-U/Qty
    lowest_col = rate_start_col + n
    lpr_col = lowest_col + 3
    incdec_col = lpr_col + 1

    banner_align = Alignment(horizontal="center", vertical="center")
    banner_font = Font(bold=True, size=10)

    title_cell = ws.cell(row=1, column=1, value=labels.cs_title)
    title_cell.font = banner_font
    title_cell.alignment = banner_align
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=incdec_col)

    inquiry_cell = ws.cell(row=2, column=1, value=cs.tender.inquiry_no)
    inquiry_cell.font = banner_font
    inquiry_cell.alignment = banner_align
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=incdec_col)

    header_row = 3
    for col, label in [(1, "Ser"), (2, "Part No"), (3, "Description"), (4, "A/U"), (5, "Qty")]:
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = bold
        cell.alignment = header_align
    ws.cell(
        row=header_row,
        column=rate_start_col,
        value=f"Rate Quoted by Firms Excl {cs.tender.tax_percent:g}% {cs.tender.tax_type.value}",
    ).font = bold
    ws.cell(row=header_row, column=lowest_col, value="Lowest").font = bold
    ws.cell(row=header_row, column=lpr_col, value="LPR (Rs)").font = bold
    ws.cell(row=header_row, column=incdec_col, value="Inc/Dec %").font = bold
    for col in (rate_start_col, lowest_col, lpr_col, incdec_col):
        ws.cell(row=header_row, column=col).alignment = header_align
    if n > 1:
        ws.merge_cells(start_row=header_row, start_column=rate_start_col, end_row=header_row, end_column=lowest_col - 1)
    ws.merge_cells(start_row=header_row, start_column=lowest_col, end_row=header_row, end_column=lowest_col + 2)

    subheader_row = header_row + 1
    for i, supplier in enumerate(suppliers):
        cell = ws.cell(row=subheader_row, column=rate_start_col + i, value=supplier.name)
        cell.font = bold
        cell.alignment = header_align
    ws.cell(row=subheader_row, column=lowest_col, value="Firm").font = bold
    ws.cell(row=subheader_row, column=lowest_col + 1, value="Rate Rs.").font = bold
    ws.cell(row=subheader_row, column=lowest_col + 2, value="Total Value").font = bold
    for col in (lowest_col, lowest_col + 1, lowest_col + 2):
        ws.cell(row=subheader_row, column=col).alignment = header_align

    # cs.item_results doesn't carry raw per-supplier rates, so pull them from
    # item.quotes (lazy-loaded via the still-open session) as we go.
    row = subheader_row + 1
    for r in cs.item_results:
        item = r.item
        ws.cell(row=row, column=1, value=item.ser)
        ws.cell(row=row, column=2, value=item.item_master.part_no)
        ws.cell(row=row, column=3, value=item.item_master.description)
        ws.cell(row=row, column=4, value=item.item_master.default_unit)
        ws.cell(row=row, column=5, value=item.qty)
        for i, supplier in enumerate(suppliers):
            rate = next((q.rate for q in item.quotes if q.supplier_id == supplier.id), None)
            ws.cell(row=row, column=rate_start_col + i, value=rate if rate is not None else "NQ")
        lowest_name = cs.suppliers_by_id[r.lowest_supplier_id].name if r.lowest_supplier_id else "NQ"
        ws.cell(row=row, column=lowest_col, value=lowest_name)
        ws.cell(row=row, column=lowest_col + 1, value=r.lowest_rate if r.lowest_rate is not None else 0)
        ws.cell(row=row, column=lowest_col + 2, value=r.total_value)
        ws.cell(row=row, column=lpr_col, value=item.lpr if item.lpr is not None else "-")
        ws.cell(row=row, column=incdec_col, value=f"{r.inc_dec_pct:.2f}" if r.inc_dec_pct is not None else "-")
        row += 1

    tax_label = cs.tender.tax_type.value
    right_align = Alignment(horizontal="right", vertical="center")

    def _total_label_row(label: str, value: float) -> None:
        nonlocal row
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = bold
        cell.alignment = right_align
        # Merged right up to the column before the value, so the
        # right-aligned label sits flush against its value - matching
        # the original CS.xlsx (merged A:J there) instead of leaving a
        # big gap between a left-aligned label and a far-right value.
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=lowest_col + 1)
        ws.cell(row=row, column=lowest_col + 2, value=value)
        row += 1

    row += 1  # blank spacer row between the item list and the totals
    _total_label_row(f"Total Amount Excl {cs.tender.tax_percent:g}% {tax_label} (Rs)", cs.grand_total.store_value)
    _total_label_row(f"{cs.tender.tax_percent:g}% {tax_label} (Rs)", cs.grand_total.tax_amount)
    _total_label_row(f"Total Amount Incl {cs.tender.tax_percent:g}% {tax_label} (Rs)", cs.grand_total.contract_value)

    row += 2
    ws.cell(row=row, column=4, value="SUMMARY").font = bold
    row += 1
    ws.cell(row=row, column=4, value="Firm").font = bold
    ws.cell(row=row, column=6, value="Items").font = bold
    ws.cell(row=row, column=7, value="Store Value").font = bold
    ws.cell(row=row, column=8, value=tax_label).font = bold
    ws.cell(row=row, column=9, value="Contr Value").font = bold
    row += 1
    for f in cs.firm_summaries:
        ws.cell(row=row, column=4, value=f.supplier_name)
        ws.cell(row=row, column=6, value=f.item_count)
        ws.cell(row=row, column=7, value=f.store_value)
        ws.cell(row=row, column=8, value=f.tax_amount)
        ws.cell(row=row, column=9, value=f.contract_value)
        row += 1
    ws.cell(row=row, column=4, value="G.Total").font = bold
    ws.cell(row=row, column=6, value=cs.grand_total.item_count).font = bold
    ws.cell(row=row, column=7, value=cs.grand_total.store_value).font = bold
    ws.cell(row=row, column=8, value=cs.grand_total.tax_amount).font = bold
    ws.cell(row=row, column=9, value=cs.grand_total.contract_value).font = bold

    # Signature/approval block, matching CS.xlsx's own multi-signatory
    # chain (Prep By -> Checked by -> HEAD QAC -> COUNTERSIGNED -> FMSAD).
    # The original stores these as floating text boxes (drawing1.xml) with
    # a specific person's name in each (e.g. "Prep By Lnk/Clk Zaheer") -
    # openpyxl can't reliably create floating shapes, and a specific name
    # doesn't belong in a reusable template anyway (same call made for the
    # PP/CA Word templates), so this reproduces it with ordinary bordered
    # cells: a blank line for the real approver to sign, with only the
    # standing role label kept, not the dummy name.
    def _sig_slot(
        line_row: int, col_start: int, col_end: int, label: str, halign: str, designation: str = ""
    ) -> None:
        bottom_border = Border(bottom=Side(style="thin"))
        for c in range(col_start, col_end + 1):
            ws.cell(row=line_row, column=c).border = bottom_border
        if col_end > col_start:
            ws.merge_cells(start_row=line_row, start_column=col_start, end_row=line_row, end_column=col_end)
        label_cell = ws.cell(row=line_row + 1, column=col_start, value=label)
        label_cell.font = Font(size=9)
        label_cell.alignment = Alignment(horizontal=halign, vertical="center")
        if col_end > col_start:
            ws.merge_cells(start_row=line_row + 1, start_column=col_start, end_row=line_row + 1, end_column=col_end)
        if designation:
            # Optional second line under the role label - e.g. "Junior
            # Clerk (BS-11)" - sourced from a Custom Field (Settings), see
            # custom_fields.SUGGESTED_CS_SIGNATURE_FIELDS. Blank/omitted if
            # that custom field was never set, matching the original
            # layout exactly.
            designation_cell = ws.cell(row=line_row + 2, column=col_start, value=designation)
            designation_cell.font = Font(size=8, italic=True)
            designation_cell.alignment = Alignment(horizontal=halign, vertical="center")
            if col_end > col_start:
                ws.merge_cells(start_row=line_row + 2, start_column=col_start, end_row=line_row + 2, end_column=col_end)

    row += 3
    right_start = max(lowest_col, incdec_col - 2)
    _sig_slot(row, 1, 3, labels.prep_by_label, halign="left", designation=custom_fields.get("prep_by_designation", ""))
    _sig_slot(
        row, right_start, incdec_col, labels.checked_by_label, halign="right",
        designation=custom_fields.get("checked_by_designation", ""),
    )

    row += 4
    _sig_slot(
        row, 1, incdec_col, labels.head_qac_label, halign="center",
        designation=custom_fields.get("head_qac_designation", ""),
    )

    row += 4
    sig_cell = ws.cell(row=row, column=1, value=labels.countersigned_label)
    sig_cell.font = Font(bold=True, size=12)
    sig_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=incdec_col)

    row += 3
    _sig_slot(
        row, 1, incdec_col, labels.fmsad_label, halign="center",
        designation=custom_fields.get("fmsad_designation", ""),
    )

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    # Dynamic columns (supplier rate columns, Lowest Firm/Rate/Total,
    # LPR, Inc/Dec%) previously had no explicit width at all, so long
    # supplier names and header labels overflowed/clipped against Excel's
    # narrow default column width - this is what looked like "overlapping"
    # or broken merges. wrap_text on the header cells (set above) lets
    # Excel grow the row height instead of needing very wide columns.
    for col in range(rate_start_col, lowest_col):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions[get_column_letter(lowest_col)].width = 18
    ws.column_dimensions[get_column_letter(lowest_col + 1)].width = 11
    ws.column_dimensions[get_column_letter(lowest_col + 2)].width = 13
    ws.column_dimensions[get_column_letter(lpr_col)].width = 11
    ws.column_dimensions[get_column_letter(incdec_col)].width = 10

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_package_cs_xlsx(cs: "ComparativeStatement", labels: DocumentLabels, custom_fields: dict = None) -> bytes:
    """Render the package-basis comparison: same item list and raw rate
    grid as export_cs_xlsx, but instead of picking the lowest rate per
    item, ranks each supplier's TOTAL across every item (cs.package_totals)
    - the whole-package-to-one-firm alternative to item-by-item awarding.
    Only a supplier who quoted every item is a real package candidate;
    others are listed but not eligible."""
    custom_fields = custom_fields or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Package Comparison"
    bold = Font(bold=True)
    header_align = Alignment(wrap_text=True, horizontal="center", vertical="center")
    lowest_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")

    suppliers = sorted(cs.suppliers_by_id.values(), key=lambda s: s.name)
    n = len(suppliers)
    rate_start_col = 6  # after Ser/Part No/Description/A-U/Qty
    last_col = max(rate_start_col + n - 1, 6)

    banner_align = Alignment(horizontal="center", vertical="center")
    banner_font = Font(bold=True, size=10)

    title_cell = ws.cell(row=1, column=1, value=f"{labels.cs_title} (PACKAGE BASIS)")
    title_cell.font = banner_font
    title_cell.alignment = banner_align
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    inquiry_cell = ws.cell(row=2, column=1, value=cs.tender.inquiry_no)
    inquiry_cell.font = banner_font
    inquiry_cell.alignment = banner_align
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)

    header_row = 3
    for col, label in [(1, "Ser"), (2, "Part No"), (3, "Description"), (4, "A/U"), (5, "Qty")]:
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = bold
        cell.alignment = header_align
    ws.cell(
        row=header_row,
        column=rate_start_col,
        value=f"Rate Quoted by Firms Excl {cs.tender.tax_percent:g}% {cs.tender.tax_type.value}",
    ).font = bold
    ws.cell(row=header_row, column=rate_start_col).alignment = header_align
    if n > 1:
        ws.merge_cells(start_row=header_row, start_column=rate_start_col, end_row=header_row, end_column=last_col)

    subheader_row = header_row + 1
    for i, supplier in enumerate(suppliers):
        cell = ws.cell(row=subheader_row, column=rate_start_col + i, value=supplier.name)
        cell.font = bold
        cell.alignment = header_align

    row = subheader_row + 1
    for r in cs.item_results:
        item = r.item
        ws.cell(row=row, column=1, value=item.ser)
        ws.cell(row=row, column=2, value=item.item_master.part_no)
        ws.cell(row=row, column=3, value=item.item_master.description)
        ws.cell(row=row, column=4, value=item.item_master.default_unit)
        ws.cell(row=row, column=5, value=item.qty)
        for i, supplier in enumerate(suppliers):
            rate = next((q.rate for q in item.quotes if q.supplier_id == supplier.id), None)
            ws.cell(row=row, column=rate_start_col + i, value=rate if rate is not None else "NQ")
        row += 1

    row += 2
    tax_label = cs.tender.tax_type.value
    heading_cell = ws.cell(row=row, column=1, value="PACKAGE TOTALS - if the entire item list were awarded to one firm")
    heading_cell.font = bold
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    row += 1

    ws.cell(row=row, column=1, value="Firm").font = bold
    ws.cell(row=row, column=3, value="Items Quoted").font = bold
    ws.cell(row=row, column=4, value="Store Value").font = bold
    ws.cell(row=row, column=5, value=tax_label).font = bold
    ws.cell(row=row, column=6, value="Contract Value").font = bold
    ws.cell(row=row, column=7, value="Eligible").font = bold
    row += 1

    lowest_eligible_marked = False
    for p in cs.package_totals:
        name_cell = ws.cell(row=row, column=1, value=p.supplier_name)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=3, value=f"{p.quoted_item_count}/{p.total_item_count}")
        ws.cell(row=row, column=4, value=p.store_value)
        ws.cell(row=row, column=5, value=p.tax_amount)
        ws.cell(row=row, column=6, value=p.contract_value)
        ws.cell(row=row, column=7, value="Yes" if p.fully_quoted else "No (partial)")
        if p.fully_quoted and not lowest_eligible_marked:
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = lowest_fill
            name_cell.font = bold
            lowest_eligible_marked = True
        row += 1

    # Signature/approval block, same convention as export_cs_xlsx.
    def _sig_slot(
        line_row: int, col_start: int, col_end: int, label: str, halign: str, designation: str = ""
    ) -> None:
        bottom_border = Border(bottom=Side(style="thin"))
        for c in range(col_start, col_end + 1):
            ws.cell(row=line_row, column=c).border = bottom_border
        if col_end > col_start:
            ws.merge_cells(start_row=line_row, start_column=col_start, end_row=line_row, end_column=col_end)
        label_cell = ws.cell(row=line_row + 1, column=col_start, value=label)
        label_cell.font = Font(size=9)
        label_cell.alignment = Alignment(horizontal=halign, vertical="center")
        if col_end > col_start:
            ws.merge_cells(start_row=line_row + 1, start_column=col_start, end_row=line_row + 1, end_column=col_end)
        if designation:
            designation_cell = ws.cell(row=line_row + 2, column=col_start, value=designation)
            designation_cell.font = Font(size=8, italic=True)
            designation_cell.alignment = Alignment(horizontal=halign, vertical="center")
            if col_end > col_start:
                ws.merge_cells(start_row=line_row + 2, start_column=col_start, end_row=line_row + 2, end_column=col_end)

    row += 3
    right_start = max(4, last_col - 2)
    _sig_slot(row, 1, 3, labels.prep_by_label, halign="left", designation=custom_fields.get("prep_by_designation", ""))
    _sig_slot(
        row, right_start, last_col, labels.checked_by_label, halign="right",
        designation=custom_fields.get("checked_by_designation", ""),
    )

    row += 4
    _sig_slot(
        row, 1, last_col, labels.head_qac_label, halign="center",
        designation=custom_fields.get("head_qac_designation", ""),
    )

    row += 4
    sig_cell = ws.cell(row=row, column=1, value=labels.countersigned_label)
    sig_cell.font = Font(bold=True, size=12)
    sig_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)

    # The item-wise export also has an FMSAD line below COUNTERSIGNED -
    # matched here for consistency, since both are "comparative statement"
    # variants meant for the same sign-off process.
    row += 3
    _sig_slot(
        row, 1, last_col, labels.fmsad_label, halign="center",
        designation=custom_fields.get("fmsad_designation", ""),
    )

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    for col in range(rate_start_col, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_working_comparison_xlsx(
    tender: Tender,
    items: List[Item],
    selected_suppliers: List[Supplier],
    view: str,
    item_results: Optional[List["ItemResult"]] = None,
    package_totals: Optional[List["PackageTotal"]] = None,
) -> bytes:
    """A lightweight, user-narrowed comparison scoped to a picked subset of
    suppliers (e.g. "lowest 5 overall", or one Sourcing Options bundle's
    members) - a working shortlist for internal review, NOT the official
    signed Comparative Statement. Deliberately shaped differently from
    export_cs_xlsx/export_package_cs_xlsx: no signature block (nothing here
    is meant to be signed/filed), no re-import support, and a banner that
    says plainly it isn't the official document, so it can't be mistaken
    for one further down the approval chain."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Working Comparison"
    bold = Font(bold=True)
    header_align = Alignment(wrap_text=True, horizontal="center", vertical="center")
    banner_align = Alignment(horizontal="center", vertical="center")
    lowest_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")

    suppliers_by_id = {s.id: s for s in selected_suppliers}
    n = len(selected_suppliers)
    rate_start_col = 5  # after Ser/Part No/Description/Qty
    if view == "package":
        last_col = max(rate_start_col + n - 1, 6)
    else:
        lowest_col = rate_start_col + n
        last_col = lowest_col + 2

    title_cell = ws.cell(row=1, column=1, value="WORKING COMPARISON - SELECTED SUPPLIERS ONLY")
    title_cell.font = Font(bold=True, size=11, color="C0392B")
    title_cell.alignment = banner_align
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    subtitle_cell = ws.cell(
        row=2, column=1,
        value="Not the official Comparative Statement - a working shortlist for internal review.",
    )
    subtitle_cell.font = Font(italic=True, size=9)
    subtitle_cell.alignment = banner_align
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)

    inquiry_cell = ws.cell(row=3, column=1, value=tender.inquiry_no)
    inquiry_cell.font = Font(bold=True, size=10)
    inquiry_cell.alignment = banner_align
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)

    suppliers_cell = ws.cell(row=4, column=1, value="Suppliers: " + ", ".join(s.name for s in selected_suppliers))
    suppliers_cell.font = Font(size=9)
    suppliers_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)

    header_row = 6
    for col, label in [(1, "Ser"), (2, "Part No"), (3, "Description"), (4, "Qty")]:
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = bold
        cell.alignment = header_align
    for i, s in enumerate(selected_suppliers):
        cell = ws.cell(row=header_row, column=rate_start_col + i, value=s.name)
        cell.font = bold
        cell.alignment = header_align

    if view == "package":
        row = header_row + 1
        for item in items:
            ws.cell(row=row, column=1, value=item.ser)
            ws.cell(row=row, column=2, value=item.item_master.part_no)
            ws.cell(row=row, column=3, value=item.item_master.description)
            ws.cell(row=row, column=4, value=item.qty)
            for i, s in enumerate(selected_suppliers):
                rate = next((q.rate for q in item.quotes if q.supplier_id == s.id), None)
                ws.cell(row=row, column=rate_start_col + i, value=rate if rate is not None else "NQ")
            row += 1

        row += 2
        tax_label = tender.tax_type.value
        heading_cell = ws.cell(row=row, column=1, value="PACKAGE TOTALS (selected suppliers only)")
        heading_cell.font = bold
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
        row += 1
        for col, label in [(1, "Firm"), (3, "Items Quoted"), (4, "Store Value"), (5, tax_label), (6, "Contract Value"), (7, "Eligible")]:
            ws.cell(row=row, column=col, value=label).font = bold
        row += 1
        lowest_marked = False
        for p in (package_totals or []):
            name_cell = ws.cell(row=row, column=1, value=p.supplier_name)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=3, value=f"{p.quoted_item_count}/{p.total_item_count}")
            ws.cell(row=row, column=4, value=p.store_value)
            ws.cell(row=row, column=5, value=p.tax_amount)
            ws.cell(row=row, column=6, value=p.contract_value)
            ws.cell(row=row, column=7, value="Yes" if p.fully_quoted else "No (partial)")
            if p.fully_quoted and not lowest_marked:
                for c in range(1, 8):
                    ws.cell(row=row, column=c).fill = lowest_fill
                name_cell.font = bold
                lowest_marked = True
            row += 1
    else:
        ws.cell(row=header_row, column=lowest_col, value="Lowest Firm").font = bold
        ws.cell(row=header_row, column=lowest_col + 1, value="Rate Rs.").font = bold
        ws.cell(row=header_row, column=lowest_col + 2, value="Total Value").font = bold
        for col in (lowest_col, lowest_col + 1, lowest_col + 2):
            ws.cell(row=header_row, column=col).alignment = header_align

        row = header_row + 1
        for r in (item_results or []):
            item = r.item
            ws.cell(row=row, column=1, value=item.ser)
            ws.cell(row=row, column=2, value=item.item_master.part_no)
            ws.cell(row=row, column=3, value=item.item_master.description)
            ws.cell(row=row, column=4, value=item.qty)
            for i, s in enumerate(selected_suppliers):
                rate = next((q.rate for q in item.quotes if q.supplier_id == s.id), None)
                ws.cell(row=row, column=rate_start_col + i, value=rate if rate is not None else "NQ")
            lowest_name = suppliers_by_id[r.lowest_supplier_id].name if r.lowest_supplier_id else "NQ"
            ws.cell(row=row, column=lowest_col, value=lowest_name)
            ws.cell(row=row, column=lowest_col + 1, value=r.lowest_rate if r.lowest_rate is not None else 0)
            ws.cell(row=row, column=lowest_col + 2, value=r.total_value)
            row += 1

        row += 2
        tax_percent = tender.tax_percent
        tax_label = tender.tax_type.value
        total_store = sum(r.total_value for r in (item_results or []))
        tax_amount = total_store * tax_percent / 100
        right_align = Alignment(horizontal="right", vertical="center")

        def _total_row(label: str, value: float) -> None:
            nonlocal row
            cell = ws.cell(row=row, column=1, value=label)
            cell.font = bold
            cell.alignment = right_align
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=lowest_col + 1)
            ws.cell(row=row, column=lowest_col + 2, value=value)
            row += 1

        _total_row(f"Total Amount Excl {tax_percent:g}% {tax_label} (Rs)", total_store)
        _total_row(f"{tax_percent:g}% {tax_label} (Rs)", tax_amount)
        _total_row(f"Total Amount Incl {tax_percent:g}% {tax_label} (Rs)", total_store + tax_amount)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 8
    for col in range(rate_start_col, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_purchase_proposal_xlsx(proposal: "PurchaseProposal") -> bytes:
    """Render a PurchaseProposal (app/award_engine.py) as an .xlsx workbook:
    one block per awarded firm, an unresolved-items block if any, and a
    grand total - for internal sign-off before contract drafts go out."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Proposal"

    bold = Font(bold=True)
    item_headers = ["Ser", "Part No", "Description", "Unit", "Qty", "Rate", "Total Value"]
    tax_label = proposal.tender.tax_type.value

    row = 1
    ws.cell(row=row, column=1, value=f"PURCHASE PROPOSAL - {proposal.tender.inquiry_no}").font = Font(
        bold=True, size=13
    )
    row += 2

    for group in proposal.firm_groups:
        ws.cell(row=row, column=1, value=f"Firm: {group.supplier_name}").font = bold
        row += 1

        for col, header in enumerate(item_headers, start=1):
            ws.cell(row=row, column=col, value=header).font = bold
        row += 1

        for ai in group.items:
            item = ai.item
            ws.cell(row=row, column=1, value=item.ser)
            ws.cell(row=row, column=2, value=item.item_master.part_no)
            ws.cell(row=row, column=3, value=item.item_master.description)
            ws.cell(row=row, column=4, value=item.item_master.default_unit)
            ws.cell(row=row, column=5, value=item.qty)
            ws.cell(row=row, column=6, value=ai.awarded_rate)
            ws.cell(row=row, column=7, value=ai.total_value)
            row += 1

        ws.cell(row=row, column=6, value="Store Value").font = bold
        ws.cell(row=row, column=7, value=group.store_value)
        row += 1
        ws.cell(row=row, column=6, value=tax_label).font = bold
        ws.cell(row=row, column=7, value=group.tax_amount)
        row += 1
        ws.cell(row=row, column=6, value="Contract Value").font = bold
        ws.cell(row=row, column=7, value=group.contract_value)
        row += 3

    if proposal.unresolved_items:
        ws.cell(row=row, column=1, value="Unresolved items (no valid award - not quoted / needs review)").font = bold
        row += 1
        for col, header in enumerate(["Ser", "Part No", "Description", "Unit", "Qty"], start=1):
            ws.cell(row=row, column=col, value=header).font = bold
        row += 1
        for item in proposal.unresolved_items:
            ws.cell(row=row, column=1, value=item.ser)
            ws.cell(row=row, column=2, value=item.item_master.part_no)
            ws.cell(row=row, column=3, value=item.item_master.description)
            ws.cell(row=row, column=4, value=item.item_master.default_unit)
            ws.cell(row=row, column=5, value=item.qty)
            row += 1
        row += 2

    ws.cell(row=row, column=1, value="GRAND TOTAL").font = bold
    ws.cell(row=row, column=2, value=f"Items: {proposal.grand_total.item_count}")
    ws.cell(row=row, column=6, value="Store Value").font = bold
    ws.cell(row=row, column=7, value=proposal.grand_total.store_value)
    row += 1
    ws.cell(row=row, column=6, value=tax_label).font = bold
    ws.cell(row=row, column=7, value=proposal.grand_total.tax_amount)
    row += 1
    ws.cell(row=row, column=6, value="Contract Value").font = bold
    ws.cell(row=row, column=7, value=proposal.grand_total.contract_value)

    for col_letter, width in zip("ABCDEFG", [6, 12, 36, 8, 8, 12, 14]):
        ws.column_dimensions[col_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
