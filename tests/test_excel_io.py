from io import BytesIO
from pathlib import Path

import pytest
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

from app.cs_engine import build_comparative_statement
from app.excel_io import (
    export_cs_xlsx,
    export_department_list_xlsx,
    export_item_catalog_xlsx,
    export_package_cs_xlsx,
    export_rfq_item_list_xlsx,
    export_supplier_list_xlsx,
    import_tender,
)
from app.models import Department, DocumentLabels, Item, ItemMaster, Quote, Supplier, TaxType, Tender

CS_XLSX_PATH = Path(__file__).resolve().parent.parent / "CS.xlsx"
DEFAULT_LABELS = DocumentLabels()


def _fresh_session() -> Session:
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    SQLModel.metadata.create_all(engine)
    return Session(engine)


def test_import_creates_one_tender_23_items_3_suppliers():
    with _fresh_session() as session:
        tender = import_tender(CS_XLSX_PATH, session)

        assert tender.id is not None
        assert tender.tax_percent == 18.0

        items = session.exec(select(Item).where(Item.tender_id == tender.id)).all()
        assert len(items) == 23  # includes Ser 1 & 21, NQ by every firm

        suppliers = session.exec(select(Supplier)).all()
        supplier_names = {s.name for s in suppliers}
        assert supplier_names == {
            "M/s Awan Tech",
            "M/s SNS Enterprises",
            "M/s Libra Enterprises",
        }

        quotes = session.exec(select(Quote)).all()
        assert len(quotes) == 23 * 3


def test_nq_cells_import_as_null_rate():
    with _fresh_session() as session:
        tender = import_tender(CS_XLSX_PATH, session)
        items = session.exec(select(Item).where(Item.tender_id == tender.id)).all()
        quotes = session.exec(select(Quote)).all()

        nq_count = sum(1 for q in quotes if q.rate is None)
        assert nq_count > 0

        # Ser 1 and Ser 21 were NQ by every firm in the fixture.
        for ser in (1, 21):
            item = next(i for i in items if i.ser == ser)
            item_quotes = [q for q in quotes if q.item_id == item.id]
            assert len(item_quotes) == 3
            assert all(q.rate is None for q in item_quotes)


def test_specific_rates_match_source_file():
    with _fresh_session() as session:
        tender = import_tender(CS_XLSX_PATH, session)
        items = session.exec(select(Item).where(Item.tender_id == tender.id)).all()
        quotes = session.exec(select(Quote)).all()
        suppliers = session.exec(select(Supplier)).all()
        supplier_by_name = {s.name: s for s in suppliers}

        # Ser 2 "Brush Brass Wire 6 Row": Awan 850, SNS 350, Libra 900.
        item2 = next(i for i in items if i.ser == 2)
        rates_by_supplier_id = {
            q.supplier_id: q.rate for q in quotes if q.item_id == item2.id
        }
        assert rates_by_supplier_id[supplier_by_name["M/s Awan Tech"].id] == 850
        assert rates_by_supplier_id[supplier_by_name["M/s SNS Enterprises"].id] == 350
        assert rates_by_supplier_id[supplier_by_name["M/s Libra Enterprises"].id] == 900


def test_reimporting_reuses_catalog_items_instead_of_duplicating():
    with _fresh_session() as session:
        tender_a = import_tender(CS_XLSX_PATH, session)
        tender_b = import_tender(CS_XLSX_PATH, session)

        catalog = session.exec(select(ItemMaster)).all()
        assert len(catalog) == 23  # not 46 - the second import reused every row

        items_a = session.exec(select(Item).where(Item.tender_id == tender_a.id)).all()
        items_b = session.exec(select(Item).where(Item.tender_id == tender_b.id)).all()
        item_masters_a = {i.item_master_id for i in items_a}
        item_masters_b = {i.item_master_id for i in items_b}
        assert item_masters_a == item_masters_b  # both tenders point at the same catalog rows

        # NIV part_no is reused by several genuinely different items - confirm
        # they stayed distinct catalog rows rather than collapsing into one.
        niv_descriptions = {
            im.description for im in catalog if im.part_no == "NIV"
        }
        assert len(niv_descriptions) >= 2


def test_exported_cs_round_trips_through_the_apps_own_importer():
    """Export must not just look right - re-importing it with our own
    import_tender() must reproduce the exact same computed CS, proving the
    exported file is genuinely CS.xlsx-shaped, not just presentable."""
    with _fresh_session() as original_session:
        original_tender = import_tender(CS_XLSX_PATH, original_session)
        cs = build_comparative_statement(original_session, original_tender.id)
        exported_bytes = export_cs_xlsx(cs, DEFAULT_LABELS)

    with _fresh_session() as reimport_session:
        reimported_tender = import_tender(BytesIO(exported_bytes), reimport_session)
        reimported_cs = build_comparative_statement(reimport_session, reimported_tender.id)

        assert len(reimported_cs.item_results) == 23

        summaries_by_name = {s.supplier_name: s for s in reimported_cs.firm_summaries}
        assert set(summaries_by_name) == {"M/s SNS Enterprises", "M/s Awan Tech"}

        sns = summaries_by_name["M/s SNS Enterprises"]
        assert sns.item_count == 10
        assert sns.store_value == pytest.approx(209655)
        assert sns.contract_value == pytest.approx(247392.90, abs=0.01)

        awan = summaries_by_name["M/s Awan Tech"]
        assert awan.item_count == 11
        assert awan.store_value == pytest.approx(211134)
        assert awan.contract_value == pytest.approx(249138.12, abs=0.01)

        assert reimported_cs.grand_total.item_count == 21
        assert reimported_cs.grand_total.store_value == pytest.approx(420789)
        assert reimported_cs.grand_total.contract_value == pytest.approx(496531.02, abs=0.01)

        # Ser 1 & 21 (NQ by every firm) must still be NQ after the round trip.
        for ser in (1, 21):
            result = next(r for r in reimported_cs.item_results if r.item.ser == ser)
            assert result.lowest_supplier_id is None


def test_pst_tender_computes_and_round_trips_correctly():
    """A PST tender's tax amount must compute the same way GST's does, and
    exporting/re-importing must preserve the PST tax type - not silently
    fall back to GST."""
    with _fresh_session() as session:
        tender = import_tender(CS_XLSX_PATH, session)
        tender.tax_type = TaxType.PST
        tender.tax_percent = 15.0
        session.add(tender)
        session.commit()

        cs = build_comparative_statement(session, tender.id)
        sns = next(s for s in cs.firm_summaries if s.supplier_name == "M/s SNS Enterprises")
        assert sns.tax_amount == pytest.approx(209655 * 0.15)
        assert sns.contract_value == pytest.approx(209655 * 1.15)

        exported_bytes = export_cs_xlsx(cs, DEFAULT_LABELS)

    with _fresh_session() as reimport_session:
        reimported_tender = import_tender(BytesIO(exported_bytes), reimport_session)
        assert reimported_tender.tax_type == TaxType.PST
        assert reimported_tender.tax_percent == pytest.approx(15.0)


def test_package_export_opens_and_lists_ranked_supplier_totals():
    from openpyxl import load_workbook

    with _fresh_session() as session:
        tender = import_tender(CS_XLSX_PATH, session)
        cs = build_comparative_statement(session, tender.id)
        exported_bytes = export_package_cs_xlsx(cs, DEFAULT_LABELS)

    wb = load_workbook(BytesIO(exported_bytes))
    ws = wb.active
    assert ws.title == "Package Comparison"

    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert any("PACKAGE BASIS" in str(v) for v in all_values)
    assert any("PACKAGE TOTALS" in str(v) for v in all_values)
    # Every supplier from the comparative statement should appear somewhere
    # in the package totals section, ranked/eligible-flagged.
    for supplier in cs.suppliers_by_id.values():
        assert supplier.name in all_values


def test_item_catalog_export_lists_part_no_description_unit():
    from openpyxl import load_workbook

    items = [
        ItemMaster(id=1, part_no="A-1", description="Widget", default_unit="Nos"),
        ItemMaster(id=2, part_no="A-2", description="Gadget", default_unit="Kg"),
    ]
    wb = load_workbook(BytesIO(export_item_catalog_xlsx(items)))
    ws = wb.active
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert "A-1" in all_values and "Widget" in all_values and "Nos" in all_values
    assert "A-2" in all_values and "Gadget" in all_values and "Kg" in all_values


def test_supplier_list_export_lists_contact_details():
    from openpyxl import load_workbook

    suppliers = [Supplier(id=1, name="M/s Test Co", phone="12345", email="a@b.com")]
    wb = load_workbook(BytesIO(export_supplier_list_xlsx(suppliers)))
    ws = wb.active
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert "M/s Test Co" in all_values
    assert "12345" in all_values
    assert "a@b.com" in all_values


def test_department_list_export_lists_names():
    from openpyxl import load_workbook

    departments = [Department(id=1, name="Finance"), Department(id=2, name="Admin")]
    wb = load_workbook(BytesIO(export_department_list_xlsx(departments)))
    ws = wb.active
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert "Finance" in all_values
    assert "Admin" in all_values


def test_rfq_item_list_export_has_no_pricing_columns():
    from openpyxl import load_workbook

    with _fresh_session() as session:
        tender = Tender(id=1, inquiry_no="RFQ-1")
        session.add(tender)
        im = ItemMaster(part_no="A-1", description="Widget", default_unit="Nos")
        session.add(im)
        session.flush()
        item = Item(tender_id=1, item_master_id=im.id, ser=1, qty=10)
        session.add(item)
        session.commit()

        items = session.exec(select(Item).where(Item.tender_id == 1)).all()
        wb = load_workbook(BytesIO(export_rfq_item_list_xlsx(tender, items)))

    ws = wb.active
    header_row = [c.value for c in ws[3]]
    assert header_row == ["Ser", "Part No", "Description", "Unit", "Qty"]
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert "RFQ-1" in "".join(str(v) for v in all_values)  # title banner has the inquiry no


def test_rfq_item_list_export_handles_slash_in_inquiry_no():
    """Real inquiry numbers look like "PROC/2026/204" - Excel forbids
    \\ / ? * [ ] : in a sheet name, so the raw inquiry_no can't be used for
    ws.title directly even though it's fine in the banner cell."""
    from openpyxl import load_workbook

    tender = Tender(id=1, inquiry_no="PROC/2026/204")
    wb = load_workbook(BytesIO(export_rfq_item_list_xlsx(tender, [])))
    ws = wb.active
    assert "/" not in ws.title
    all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert any("PROC/2026/204" in str(v) for v in all_values)


def test_custom_labels_appear_in_both_cs_exports_instead_of_defaults():
    from openpyxl import load_workbook

    custom = DocumentLabels(
        cs_title="CUSTOM TITLE BANNER",
        prep_by_label="Drafted by",
        checked_by_label="Reviewed by",
        head_qac_label="Custom Head Role",
        countersigned_label="CUSTOM COUNTERSIGN",
        fmsad_label="Custom Final Approver",
    )
    with _fresh_session() as session:
        tender = import_tender(CS_XLSX_PATH, session)
        cs = build_comparative_statement(session, tender.id)

        item_wise_values = [
            c.value for row in load_workbook(BytesIO(export_cs_xlsx(cs, custom))).active.iter_rows() for c in row
        ]
        package_values = [
            c.value
            for row in load_workbook(BytesIO(export_package_cs_xlsx(cs, custom))).active.iter_rows()
            for c in row
        ]

    for values in (item_wise_values, package_values):
        assert any("CUSTOM TITLE BANNER" in str(v) for v in values if v)
        assert "Drafted by" in values
        assert "Reviewed by" in values
        assert "Custom Head Role" in values
        assert "CUSTOM COUNTERSIGN" in values
        assert "Custom Final Approver" in values
        # None of the old hardcoded defaults should leak through.
        assert "Prep By" not in values
        assert "FMSAD (XDS)" not in values
