from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

from app.db import get_session
from app.document_labels import get_document_labels
from app.main import app
from app.models import DocumentLabels

try:
    from fastapi.testclient import TestClient
except ImportError:  # pragma: no cover
    TestClient = None


def _fresh_session() -> Session:
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    SQLModel.metadata.create_all(engine)
    return Session(engine)


def _make_client():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    SQLModel.metadata.create_all(engine)

    def override_get_session():
        with Session(engine) as session:
            yield session

    app.dependency_overrides[get_session] = override_get_session
    return TestClient(app), engine


def test_get_document_labels_creates_default_singleton_once():
    with _fresh_session() as session:
        labels1 = get_document_labels(session)
        assert labels1.id == 1
        # Defaults match the strings that used to be hardcoded in excel_io.py.
        assert labels1.cs_title == "COMPARATIVE STATEMENT"
        assert labels1.prep_by_label == "Prep By"
        assert labels1.fmsad_label == "FMSAD (XDS)"

        labels2 = get_document_labels(session)
        assert labels2.id == labels1.id
        assert len(session.exec(select(DocumentLabels)).all()) == 1


def test_settings_page_updates_persist_and_affect_next_cs_export():
    client, engine = _make_client()
    try:
        resp = client.post("/tenders", data={"inquiry_no": "Labels Test", "tax_percent": "10"}, follow_redirects=False)
        tender_id = int(resp.headers["location"].rsplit("/", 1)[-1])

        # Default label first.
        resp = client.get(f"/tenders/{tender_id}/export")
        assert resp.status_code == 200
        values = [c.value for row in load_workbook(BytesIO(resp.content)).active.iter_rows() for c in row]
        assert "Prep By" in values

        # Change it via the real settings form.
        resp = client.post(
            "/settings/cs-labels",
            data={
                "cs_title": "COMPARATIVE STATEMENT",
                "prep_by_label": "Compiled by",
                "checked_by_label": "Checked by",
                "head_qac_label": "HEAD QAC (TDA)",
                "countersigned_label": "COUNTERSIGNED",
                "fmsad_label": "FMSAD (XDS)",
            },
            follow_redirects=False,
        )
        assert resp.status_code == 303

        resp = client.get(f"/tenders/{tender_id}/export")
        values = [c.value for row in load_workbook(BytesIO(resp.content)).active.iter_rows() for c in row]
        assert "Compiled by" in values
        assert "Prep By" not in values
    finally:
        app.dependency_overrides.clear()
